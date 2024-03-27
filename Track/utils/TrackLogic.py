import slicer
# from slicer import ScriptedLoadableModuleLogic
from slicer.ScriptedLoadableModule import *
import qt, vtk, ctk

import os, csv, re

class TrackLogic(ScriptedLoadableModuleLogic):
  """This class should implement all the actual
  computation done by your module.  The interface
  should be such that other python code can import
  this class and make use of the functionality without
  requiring an instance of the Widget.
  Uses ScriptedLoadableModuleLogic base class, available at:
  https://github.com/Slicer/Slicer/blob/master/Base/Python/slicer/ScriptedLoadableModule.py
  """

  def __init__(self):
    """
    Called when the logic class is instantiated. Can be used for initializing member variables.
    """
    ScriptedLoadableModuleLogic.__init__(self)
    self.timer = qt.QTimer()
    self.redBackground = None
    self.greenBackground = None
    self.yellowBackground = None
    self.backgrounds = {
      "Red": self.redBackground,
      "Green": self.greenBackground,
      "Yellow": self.yellowBackground
    }

  def setDefaultParameters(self, customParameterNode):
    """
    Initialize parameter node with default settings.
    """
    customParameterNode.totalImages = 0
    customParameterNode.fps = 5.0  # frames (i.e. images) per second
    customParameterNode.opacity = 1.0  # 100 %
    customParameterNode.overlayAsOutline = True

  def loadImagesIntoSequenceNode(self, shNode, path):
    """
    Loads the cine images located within the provided path into 3D Slicer. They are
    placed within a sequence node and the loaded image nodes are deleted thereafter.
    :param shNode: node representing the subject hierarchy
    :param path: path to folder containing the 2D images to be imported
    """
    # NOTE: This represents a node within the MRML scene, not within the subject hierarchy
    imagesSequenceNode = None

    # Find all the image file names within the provided dir
    imageFiles = []
    for item in os.listdir(path):
      if re.match('.*\.(mha|dcm|nrrd|nii|img|nhdr)', item): # Only look for valid files
        imageFiles.append(item)
    imageFiles.sort()

    # We only want to create a sequence node if image files were found within the provided path
    if len(imageFiles) != 0:
      imagesSequenceNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSequenceNode",
                                                              "Image Nodes Sequence")

      # Create a progress/loading bar to display the progress of the images loading process
      progressDialog = qt.QProgressDialog("Loading cine images", "Cancel",
                                          0, len(imageFiles))
      progressDialog.minimumDuration = 0

      for fileIndex in range(len(imageFiles)):
        # If the 'Cancel' button was pressed, we want to return to a default state
        if progressDialog.wasCanceled:
          # Remove sequence node
          slicer.mrmlScene.RemoveNode(imagesSequenceNode)
          return None, True

        filepath = os.path.join(path, imageFiles[fileIndex])
        nodeName = (f"Image {fileIndex + 1} ({imageFiles[fileIndex]})").format(filepath)

        loadedImageNode = slicer.util.loadVolume(filepath, {"singleFile": True, "show": False})
        loadedImageNode.SetName(nodeName)
        # Place image node into sequence
        imagesSequenceNode.SetDataNodeAtValue(loadedImageNode, str(fileIndex))
        # Remove loaded image node
        imageID = shNode.GetItemByDataNode(loadedImageNode)
        shNode.RemoveItem(imageID)

        #  Update how far we are in the progress bar
        progressDialog.setValue(fileIndex + 1)

        # This render step is needed for the progress bar to visually update in the GUI
        slicer.util.forceRenderAllViews()
        slicer.app.processEvents()

      print(f"{len(imageFiles)} cine images were loaded into 3D Slicer")

      # We do the following to clear the view of the slices. I expected {"show": False} to
      # prevent anything from being shown at all, but the first loaded image will appear in the
      # foreground. This seems to be a bug in 3D Slicer.
      self.clearSliceForegrounds()

    return imagesSequenceNode, False

  def getColumnNamesFromTransformsInput(self, filepath):
      
    fileName = os.path.basename(filepath)
    fileExtension = os.path.splitext(filepath)[1]

    if re.match('.*\.(csv|xls|xlsx|txt)', filepath):
      # Check that the transforms file is a .csv type
      if filepath.endswith('.csv'):
        encodings = ["utf-8-sig", "cp1252", "iso-8859-1", "latin1"]
        for encoding in encodings:
          try:
            with open(filepath, "r", encoding = encoding) as f:
              reader = csv.reader(f)
              headers = next(reader)
              # if we can read without error, break of the encoding loop
              break
          except:
            print(f"Encoding {encoding} failed, trying next encoding")
        return headers
    # TODO - add support for .txt, xls and .xlsx files       
      if filepath.endswith('.txt'):
        with open(filepath, "r") as f:
          headers = next(f).strip().split(',')
          return headers
      if filepath.endswith('.xlsx'):
        try:
          global openpyxl
          import openpyxl
        except ModuleNotFoundError:
          if slicer.util.confirmOkCancelDisplay(f"To load {fileName}, install the 'xlrd' Python package. Click OK to install now."):
            try:
              # Create a loading popup
              messageBox = qt.QMessageBox()
              messageBox.setIcon(qt.QMessageBox.Information)
              messageBox.setWindowTitle("Package Installation")
              messageBox.setText("Installing 'openpyxl' package...")
              messageBox.setStandardButtons(qt.QMessageBox.NoButton)
              messageBox.show()
              slicer.app.processEvents()
              
              slicer.util.pip_install('openpyxl')
              import openpyxl
              
              messageBox.setText(f"'openpyxl' package installed successfully. {fileName} will now load.")
              slicer.app.processEvents()  # Process events to allow the dialog to update
              qt.QTimer.singleShot(3000, messageBox.accept)

              # Wait for user interaction
              while messageBox.isVisible():
                slicer.app.processEvents()

              messageBox.hide()  # Hide the message box
              
            except:
              slicer.util.warningDisplay(f"{fileName} file failed to load.\nPlease load a .csv or .txt file instead. ",
                                        "Failed to Load File")
              return
          else:
            slicer.util.warningDisplay(f"{fileName} failed to load.\nPlease load a .csv or .txt file instead. ",
                                       "Failed to Load File")
            return
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        headers = next(sheet.iter_rows(values_only=True))
        return headers
      elif filepath.endswith('.xls'):
        try:
          global xlrd
          import xlrd
        except ModuleNotFoundError:
          if slicer.util.confirmOkCancelDisplay(f"To load {fileName}, install the 'xlrd' Python package. Click OK to install now."):
            try:
              # Create a loading popup
              messageBox = qt.QMessageBox()
              messageBox.setIcon(qt.QMessageBox.Information)
              messageBox.setWindowTitle("Package Installation")
              messageBox.setText("Installing 'xlrd' package...")
              messageBox.setStandardButtons(qt.QMessageBox.NoButton)
              messageBox.show()
              slicer.app.processEvents()
              
              slicer.util.pip_install('xlrd')
              import xlrd
              
              messageBox.setText(f"'xlrd' package installed successfully. {fileName} will now load.")
              slicer.app.processEvents()  # Process events to allow the dialog to update
              qt.QTimer.singleShot(3000, messageBox.accept)

              # Wait for user interaction
              while messageBox.isVisible():
                slicer.app.processEvents()

              messageBox.hide()  # Hide the message box
              
            except:
              slicer.util.warningDisplay(f"{fileName} file failed to load.\nPlease load a .csv or .txt file instead. ",
                                        "Failed to Load File")
              return
          else:
            slicer.util.warningDisplay(f"{fileName} failed to load.\nPlease load a .csv or .txt file instead. ",
                                      "Failed to Load File")
            return
        wb = xlrd.open_workbook(filepath)
        sheet = wb.sheet_by_index(0)
        print(sheet.row(0))
        return sheet.row_values(0)
    
    # if we get here, we failed to read the the headers -> print out warning and return a empty list for headers   
    slicer.util.warningDisplay(f"Cannot read header row from {fileName}.\nPlease load another file instead. ",
                                  "Failed to Load File")
    return []

  def validateTransformsInput(self, filepath, numImages,headers):
    """
    Checks to ensure that the data in the provided transformation file is valid and matches the
    number of 2D images that have been loaded into 3D Slicer.
    :param filepath: path to the transforms file (which should be a .csv file)
    :param numImages: the number of cine images that have already been loaded
    """
    # NOTE: The current logic of this function will only ensure that the first {numImages}
    # transformations found within the CSV file are valid, so playback can occur. The playback will
    # still occur if later transformations after the first {numImages} transformations are corrupt.
    transformationsList = []
    fileName = os.path.basename(filepath)
    fileExtension = os.path.splitext(filepath)[1]
    headerX = headers[0]
    headerY = headers[1]
    headerZ = headers[2]
    if re.match('.*\.(csv|xls|xlsx|txt)', filepath):
      # Check that the transforms file is a .csv type
      if filepath.endswith('.csv'):
        encodings = ["utf-8-sig", "cp1252", "iso-8859-1", "latin1"]
        for encoding in encodings:
          try:
            with open(filepath, "r", encoding = encoding) as f:
              # Using a DictReader allows us to recognize the CSV header
              reader = csv.DictReader(f)
              for row in reader:
                # Extract floating point values from row
                transformationsList.append([float(row[headerX]), float(row[headerY]), float(row[headerZ])])
              
              # if we can read the file without error, break the encoding loop
              break
          except:
            print(f"Encoding {encoding} failed, trying next encoding")
          
      if len(transformationsList) == 0 and filepath.endswith('.csv'):
        slicer.util.warningDisplay(f"{fileName} file failed to load.\nPlease load another file instead. ",
                                  "Failed to Load File")
        return
      
      # TODO - add support for column selectors for .txt, xls and .xlsx files        
      # Check that the transforms file is a .txt type
      elif filepath.endswith('.txt'):
        with open(filepath, "r") as f:
          next(f)
          for line in f:
            values = line.strip().split(',')
            try:
              x, y, z = map(float, values)
              transformationsList.append([x, y, z])
            except:
              # If there was an error reading the values, break out because we can't/shouldn't
              # perform the playback if the transformation data is corrupt or missing.
              slicer.util.warningDisplay(f"An error was encountered while reading the {fileExtension} file: "
                                   f"{fileName}",
                                   "Validation Error")
              break

      # Check that the transforms file is a .xlsx type
      elif filepath.endswith('.xlsx'):         
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        rows = iter(sheet.iter_rows(values_only=True))
        next(rows)  # Start from the second row, assuming first row is header
        for row in rows:
          try:
            [x, y, z] = row
            transformationsList.append([x,y,z])
          except:
            slicer.util.warningDisplay(f"{fileName} file failed to load.\nPlease load a .csv or .txt file instead. ",
                                      "Failed to Load File")

            break
        
      # Check that the transforms file is a .xls type
      elif filepath.endswith('.xls'):            
        workbook = xlrd.open_workbook(filepath)
        sheet = workbook.sheet_by_index(0)
        for row_idx in range(1, sheet.nrows):  # Start from the second row, assuming first row is header
          values = sheet.row_values(row_idx)
          try:
            x, y, z = map(float, values)
            transformationsList.append([x, y, z])
          except:
            # If there was an error reading the values, break out because we can't/shouldn't
            # perform the playback if the transformation data is corrupt or missing.
            slicer.util.warningDisplay(f"An error was encountered while reading the {fileExtension} file: "
                                     f"{fileName}",
                                     "Validation Error")
            break


      if len(transformationsList) == numImages:
        return transformationsList
      else:
        # Extension will not create transforms nodes if the number of cine images and
        # the number of rows in the transforms file are not equal
        print(os.path.basename(filepath))
        slicer.util.warningDisplay(f"Error loading transforms file. Ensure proper formatting and matching number of transforms to cine images",
                           "Validation Error")
        
        return None

  def createTransformNodesFromTransformData(self, shNode, transforms, numImages):
    """
    For every image and it's matching transformation, create a transform node which will hold
    the transformation data for that image wthin 3D Slicer. Place them in a sequence node.
    :param shNode: node representing the subject hierarchy
    :param transforms: list of transforms extrapolated from the transforms .csv file
    :param numImages: number of 2D images loaded into 3D Slicer
    """
    # NOTE: This represents a node within the MRML scene, not within the subject hierarchy
    transformsSequenceNode = slicer.mrmlScene.AddNewNodeByClass("vtkMRMLSequenceNode",
                                                                "Transform Nodes Sequence")

    # Create a progress/loading bar to display the progress of the node creation process
    progressDialog = qt.QProgressDialog("Creating Transform Nodes From Transformation Data", "Cancel",
                                        0, numImages)
    progressDialog.minimumDuration = 0

    # 3D Slicer works with 4x4 transform matrices internally
    LPSToRASMatrix = vtk.vtkMatrix4x4()
    LPSToRASMatrix.SetElement(0, 0, -1)
    LPSToRASMatrix.SetElement(1, 1, -1)

    # NOTE: It is very important that we loop using the number of 2D images loaded, versus the size
    # of the transforms array/list. This is because we may provide a CSV with more transforms than
    # needed, but we only need to create as many transform nodes as there are 2D images.
    for i in range(numImages):
      # If the 'Cancel' button was pressed, we want to return to a default state
      if progressDialog.wasCanceled:
        # Remove sequence node
        shNode.RemoveNode(transformsSequenceNode)
        return None

      # 3D Slicer uses the RAS (Right, Anterior, Superior) basis for their coordinate system.
      # However, the transformation data we use was generated outside of 3D Slicer, using DICOM
      # images, which corresponds to the LPS (Left, Prosterier, Superior) basis. In order to use
      # this data, we must convert it from LPS to RAS, in order to correctly transform the images
      # we load into 3D Slicer. See the following links for more detail:
      # https://www.slicer.org/wiki/Coordinate_systems#Anatomical_coordinate_system
      # https://github.com/Slicer/Slicer/blob/main/Libs/MRML/Core/vtkITKTransformConverter.h#L246
      # This is a simple conversion. It can be mathematically represented as:
      # /ΔLR\   /-1  0  0  0\   /X\
      # |ΔPA| = | 0 -1  0  0| * |Y|
      # |ΔIS|   | 0  0  1  0|   |Z|
      # \ 0 /   \ 0  0  0  1/   \0/
      # Where X, Y, and Z represent the transformation in LPS.

      # Convert transform from LPS to RAS
      currentTransform = transforms[i]
      currentTransform.append(0) # Needs to be 4x1 to multiply with a 4x4
      convertedTransform = [0, 0, 0, 0]
      LPSToRASMatrix.MultiplyPoint(currentTransform, convertedTransform)

      # Create a transform matrix from the converted transform
      transformMatrix = vtk.vtkMatrix4x4()
      transformMatrix.SetElement(0, 3, convertedTransform[0])  # LR translation
      transformMatrix.SetElement(1, 3, convertedTransform[1])  # PA translation
      transformMatrix.SetElement(2, 3, convertedTransform[2])  # IS translation

      # Create a LinearTransform node to hold our transform matrix
      transformNode = \
             slicer.mrmlScene.AddNewNodeByClass("vtkMRMLLinearTransformNode", f"Transform {i + 1}")
      transformNode.ApplyTransformMatrix(transformMatrix)

      # Add the transform node to the transforms sequence node
      transformsSequenceNode.SetDataNodeAtValue(transformNode, str(i))
      # Remove the transform node
      transformNodeID = shNode.GetItemByDataNode(transformNode)
      shNode.RemoveItem(transformNodeID)

      # Update how far we are in the progress bar
      progressDialog.setValue(i + 1)

      # This render step is needed for the progress bar to visually update in the GUI
      slicer.util.forceRenderAllViews()
      slicer.app.processEvents()

    print(f"{numImages} transforms were loaded into 3D Slicer as transform nodes")
    return transformsSequenceNode

  def clearSliceForegrounds(self):
    """
    Clear each slice view from having anything visible in the foreground. This often happens
    inadvertently when using loadVolume() with "show" set to False.
    """
    layoutManager = slicer.app.layoutManager()
    for viewName in layoutManager.sliceViewNames():
      layoutManager.sliceWidget(viewName).mrmlSliceCompositeNode().SetForegroundVolumeID("None")

  def visualize(self, sequenceBrowser, sequenceNode2DImages, segmentationLabelMapID,
                    sequenceNodeTransforms, opacity, overlayAsOutline):
    """
    Visualizes the image data (2D images and 3D segmentation overlay) within the slice views and
    enables the alignment of the 3D segmentation label map according to the transformation data.
    :param sequenceBrowser: sequence browser node used to control the playback operation
    :param sequenceNode2DImages: sequence node containing the 2D images
    :param segmentationLabelMapID: subject hierarchy ID of the 3D segmentation label map
    :param sequenceNodeTransforms: sequence node containing the transforms
    :param opacity: opacity value of overlay layer (3D segmentation label map layer)
    :param overlayAsOutline: whether to show the overlay as an outline or a filled region
    """
    shNode = slicer.mrmlScene.GetSubjectHierarchyNode()
    layoutManager = slicer.app.layoutManager()

    # The proxy image node represents the current selected image within the sequence
    proxy2DImageNode = sequenceBrowser.GetProxyNode(sequenceNode2DImages)
    # The proxy transform node represents the current selected transform within the sequence
    proxyTransformNode = sequenceBrowser.GetProxyNode(sequenceNodeTransforms)
    labelMapNode = shNode.GetItemDataNode(segmentationLabelMapID)
    
    sliceWidget = self.getSliceWidget(layoutManager, proxy2DImageNode)

    name = None
    fitSlice = None
    if sliceWidget is not None:
      name = sliceWidget.sliceViewName
        
      sliceCompositeNode = sliceWidget.mrmlSliceCompositeNode()

      volumesLogic = slicer.modules.volumes.logic()
      
      # Checks if the current slice node is not showing an image
      fitSlice = False
      if sliceCompositeNode.GetLabelVolumeID() is None:
        fitSlice = True
      
      sliceCompositeNode.SetLabelVolumeID(labelMapNode.GetID())
      sliceCompositeNode.SetLabelOpacity(opacity)
      
      sliceCompositeNode.SetBackgroundVolumeID(labelMapNode.GetID())
      
      # Get the current slice node
      sliceNode = sliceWidget.mrmlSliceNode()

      # Display the label map overlay as an outline
      sliceNode.SetUseLabelOutline(overlayAsOutline)

      # Set the background volume for the current slice view
      sliceCompositeNode.SetBackgroundVolumeID(proxy2DImageNode.GetID())

      # Translate the 3D segmentation label map using the transform data
      if proxyTransformNode is not None:
        labelMapNode.SetAndObserveTransformNodeID(proxyTransformNode.GetID())
      
      sliceNode.SetSliceVisible(True)

    # Make the 3D segmentation visible in the 3D view
    tmpIdList = vtk.vtkIdList() # The nodes you want to display need to be in a vtkIdList
    tmpIdList.InsertNextId(segmentationLabelMapID)
    threeDViewNode = layoutManager.activeMRMLThreeDViewNode()
    shNode.ShowItemsInView(tmpIdList, threeDViewNode)

    # If the sliceNode is now showing an image, fit the slice view to the current background image   
    if fitSlice:
      sliceWidget.fitSliceToBackground()
    
    # Preserve previous slices
    # If a background node for the specified orientation exists, update it with the current slice
    # Otherwise, create a new background node and set it as the background for the specified orientation
    
    if name in self.backgrounds:
      background = getattr(self, name.lower() + 'Background')
      if background is None:
        # Create a new background node for the orientation
        setattr(self, name.lower() + 'Background', volumesLogic.CloneVolume(slicer.mrmlScene,
                proxy2DImageNode, f"{proxy2DImageNode.GetAttribute('Sequences.BaseName')}"))
      else:
        # Background exists, just replace the data to represent the next image in the sequence
        background.SetAndObserveImageData(proxy2DImageNode.GetImageData())
        background.SetAttribute("Sequences.BaseName", proxy2DImageNode.GetAttribute("Sequences.BaseName"))
    
    # Add the image name to the slice view background variable
    if name is not None:
      currentSlice = getattr(self, name.lower() + 'Background')
      currentSlice.SetName(proxy2DImageNode.GetAttribute('Sequences.BaseName'))

    # Set the background volumes for each orientation, if they exist
    for color in self.backgrounds:
      sliceViewWindow = slicer.app.layoutManager().sliceWidget(color).sliceView()
      sliceViewWindow.cornerAnnotation().RemoveAllObservers()
      currentSlice = getattr(self, color.lower() + 'Background')
      sliceViewWindow.cornerAnnotation().ClearAllTexts()
      # Add desired text to slice views that have a background node
      if currentSlice is not None:
        slicer.mrmlScene.GetNodeByID(f"vtkMRMLSliceCompositeNode{color}").SetBackgroundVolumeID(currentSlice.GetID())
        imageFile = slicer.mrmlScene.GetNodeByID(f"vtkMRMLSliceCompositeNode{color}").GetNodeReference('backgroundVolume') is not None
        if imageFile:
          imageFileNameText = slicer.mrmlScene.GetNodeByID(f"vtkMRMLSliceCompositeNode{color}").GetNodeReference('backgroundVolume').GetAttribute('Sequences.BaseName')
          # Place "Current Alignment" text in the slice view corner
          sliceViewWindow = slicer.app.layoutManager().sliceWidget(color).sliceView()
          sliceViewWindow.cornerAnnotation().SetText(0, imageFileNameText)
    
    for color in self.backgrounds:
      sliceViewWindow = slicer.app.layoutManager().sliceWidget(color).sliceView()
      if sliceViewWindow.cornerAnnotation().HasObserver(vtk.vtkCommand.ModifiedEvent):
        sliceViewWindow.cornerAnnotation().RemoveAllObservers()
      if sliceViewWindow.cornerAnnotation().HasObserver(vtk.vtkCornerAnnotation.UpperLeft):
        sliceViewWindow.cornerAnnotation().RemoveAllObservers()
    
    if sliceWidget is not None:
      sliceView = sliceWidget.sliceView()
      sliceView.cornerAnnotation().SetText(vtk.vtkCornerAnnotation.UpperLeft, "Current Alignment")
    # Enable alignment of the 3D segmentation label map according to the transform data so that
    # the 3D segmentation label map overlays upon the ROI of the 2D images
    if proxyTransformNode is not None:
      labelMapNode.SetAndObserveTransformNodeID(proxyTransformNode.GetID())

    # Render changes
    slicer.util.forceRenderAllViews()
    slicer.app.processEvents()

  def getSliceWidget(self, layoutManager, imageNode):
    """
    This function helps to determine the slice widget that corresponds to the orientation of the
    provided image. (i.e. the slice widget that would display the image)
    :param layoutManager: node representing the MRML layout manager
    :param imageNode: node representing the 2D image
    """
    # Determine the orientation of the image
    tmpMatrix = vtk.vtkMatrix4x4()
    if imageNode is not None:
      imageNode.GetIJKToRASMatrix(tmpMatrix)
      scanOrder = imageNode.ComputeScanOrderFromIJKToRAS(tmpMatrix)

      if scanOrder == "LR" or scanOrder == "RL":
        imageOrientation = "Sagittal"
      elif scanOrder == "AP" or scanOrder == "PA":
        imageOrientation = "Coronal"
      elif scanOrder == "IS" or scanOrder == "SI":
        imageOrientation = "Axial"
      else:
        print(f"Error: Unexpected image scan order {scanOrder}.")
        exit(1)

      # Find the slice widget that has the same orientation as the image
      sliceWidget = None
      for name in layoutManager.sliceViewNames():
        if layoutManager.sliceWidget(name).sliceOrientation == imageOrientation:
          sliceWidget = layoutManager.sliceWidget(name)

      if not sliceWidget:
        print(f"Error: A slice with the {imageOrientation} orientation was not found.")
        exit(1)

      return sliceWidget